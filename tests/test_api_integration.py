"""End-to-end API integration tests (upload, profile, drift, report, analytics)."""

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from api import app

client = TestClient(app)

RETAIL_CLEAN = b"""store_id,date,sku,sales
1,2024-06-01,A,100
2,2024-06-02,B,200
3,2024-06-03,C,150
"""

RETAIL_WITH_NULL = b"""store_id,date,sku,sales
1,2024-06-01,A,100
2,,B,200
3,2024-06-03,C,-5
"""

BASELINE_VALUE = b"value\n" + b"\n".join(str(i).encode() for i in range(50))
SHIFTED_VALUE = b"value\n" + b"\n".join(str(i + 100).encode() for i in range(50))


def _upload(
    content: bytes,
    *,
    filename: str = "test.csv",
    profile: str = "generic",
    required: str | None = None,
    baseline: bytes | None = None,
    baseline_name: str = "baseline.csv",
    sheet_name: str | None = None,
):
    files = {"file": (filename, io.BytesIO(content), "application/octet-stream")}
    data: dict[str, str] = {"quality_profile": profile}
    if required:
        data["required_columns"] = required
    if sheet_name:
        data["sheet_name"] = sheet_name
    if baseline is not None:
        files["baseline"] = (baseline_name, io.BytesIO(baseline), "application/octet-stream")
    return client.post("/api/upload", files=files, data=data)


def _make_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["store_id", "date", "sku", "sales"])
    ws.append([1, "2024-06-01", "A", 100])
    ws.append([2, "2024-06-02", "B", 200])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_formats_list():
    r = client.get("/api/formats")
    assert r.status_code == 200
    exts = {f["extension"] for f in r.json()["formats"]}
    assert ".xlsx" in exts
    assert ".json" in exts
    r = client.get("/api/health")
    assert r.status_code == 200
    assert r.json()["status"] == "ok"


def test_profiles_list():
    r = client.get("/api/profiles")
    assert r.status_code == 200
    ids = {p["id"] for p in r.json()["profiles"]}
    assert {"generic", "retail", "healthcare", "financial", "survey", "ml_training"} <= ids


def test_retail_upload_contract_passes():
    r = _upload(RETAIL_CLEAN, profile="retail", required="store_id,date,sku,sales")
    assert r.status_code == 200, r.text
    body = r.json()
    pa = body["profile_assessment"]
    assert pa["profile_id"] == "retail"
    assert pa["contract_passed"] is True
    assert pa["overall"] >= 0
    assert body["quality_profile_id"] == "retail"
    assert body["analytics"] is not None
    assert "row_completeness" in body["analytics"]
    assert body["column_count"] == 4


def test_retail_upload_contract_fails_on_null_and_negative_sales():
    r = _upload(
        RETAIL_WITH_NULL,
        profile="retail",
        required="store_id,date,sku,sales",
    )
    assert r.status_code == 200
    pa = r.json()["profile_assessment"]
    assert pa["contract_passed"] is False
    failed = [rule for rule in pa["rules"] if not rule["passed"]]
    assert len(failed) >= 1


def test_baseline_drift_distribution_shift():
    r = _upload(
        SHIFTED_VALUE,
        filename="current.csv",
        baseline=BASELINE_VALUE,
    )
    assert r.status_code == 200
    drift = r.json()["schema_drift"]
    assert drift is not None
    assert len(drift["distribution_shifted"]) >= 1


def test_report_includes_profile_assessment():
    r = _upload(RETAIL_CLEAN, profile="retail", required="store_id,date,sku,sales")
    session_id = r.json()["session_id"]
    report = client.get(f"/api/session/{session_id}/report")
    assert report.status_code == 200
    md = report.json()["markdown"]
    assert "Profile Score" in md
    assert "Legacy profiler score" in md
    assert "store_id" in md or "Column Overview" in md


def test_column_histogram_for_numeric():
    r = _upload(RETAIL_CLEAN, profile="retail")
    session_id = r.json()["session_id"]
    col = client.get(f"/api/session/{session_id}/column/sales")
    assert col.status_code == 200
    data = col.json()
    assert data["profile"]["name"] == "sales"
    assert data["histogram"] is not None
    assert len(data["histogram"]["bins"]) == len(data["histogram"]["counts"])


def test_row_completeness_nan_string():
    csv = b"a,b\nok,x\nnan,y\n"
    r = _upload(csv)
    assert r.status_code == 200
    hist = r.json()["analytics"]["row_completeness"]
    one_null = next(row for row in hist if row["nulls_in_row"] == 1)
    assert one_null["row_count"] == 1


def test_upload_returns_400_on_invalid_file():
    r = client.post(
        "/api/upload",
        files={"file": ("bad.pdf", io.BytesIO(b"%PDF"), "application/pdf")},
    )
    assert r.status_code == 400


def test_xlsx_upload_retail_profile():
    r = _upload(_make_xlsx_bytes(), filename="retail.xlsx", profile="retail")
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["column_count"] == 4
    assert body["profile_assessment"]["profile_id"] == "retail"


def _make_multi_sheet_xlsx() -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Stores"
    ws1.append(["store_id", "sales"])
    ws1.append(["1", "100"])
    ws2 = wb.create_sheet("Archive")
    ws2.append(["store_id", "sales"])
    ws2.append(["9", "5"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_inspect_excel_sheets():
    r = client.post(
        "/api/inspect",
        files={
            "file": (
                "multi.xlsx",
                io.BytesIO(_make_multi_sheet_xlsx()),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 200
    body = r.json()
    assert body["has_sheets"] is True
    assert body["sheets"] == ["Stores", "Archive"]


def test_upload_excel_sheet_by_name():
    r = _upload(
        _make_multi_sheet_xlsx(),
        filename="multi.xlsx",
        sheet_name="Archive",
    )
    assert r.status_code == 200
    assert r.json()["row_count"] == 1
    assert r.json()["sheet_name"] == "Archive"


def test_delete_session():
    sid = _upload(RETAIL_CLEAN).json()["session_id"]
    r = client.delete(f"/api/session/{sid}")
    assert r.status_code == 200
    assert r.json()["ok"] is True
    assert client.get(f"/api/session/{sid}").status_code == 404


def test_revision_increments_on_fixes():
    sid = _upload(RETAIL_CLEAN).json()["session_id"]
    initial = client.get(f"/api/session/{sid}").json()["revision"]
    r = client.post(
        "/api/fixes",
        json={"session_id": sid, "fixes": {"sales": "strip_whitespace"}},
    )
    assert r.status_code == 200
    assert r.json()["revision"] == initial + 1
