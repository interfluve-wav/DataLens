"""Tests for multi-format tabular loading."""

import io
import json

import pandas as pd
import pytest
from openpyxl import Workbook

from tabular_loader import (
    is_supported_filename,
    load_tabular,
    normalize_dataframe,
)


def test_is_supported_filename():
    assert is_supported_filename("data.csv")
    assert is_supported_filename("report.XLSX")
    assert is_supported_filename("export.json")
    assert not is_supported_filename("image.png")
    assert not is_supported_filename("")


def test_load_json_array():
    payload = [{"a": 1, "b": "x"}, {"a": 2, "b": "y"}]
    raw = json.dumps(payload).encode()
    df = load_tabular(raw, "rows.json")
    assert list(df.columns) == ["a", "b"]
    assert len(df) == 2
    assert df.loc[0, "a"] == "1"


def test_load_tsv():
    raw = b"name\tscore\nalice\t10\nbob\t20\n"
    df = load_tabular(raw, "data.tsv")
    assert list(df.columns) == ["name", "score"]
    assert len(df) == 2


def _make_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["store_id", "date", "sku", "sales"])
    ws.append([1, "2024-06-01", "A", 100])
    ws.append([2, "2024-06-02", "B", 200])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_load_xlsx():
    df = load_tabular(_make_xlsx_bytes(), "retail.xlsx")
    assert list(df.columns) == ["store_id", "date", "sku", "sales"]
    assert len(df) == 2
    assert df.loc[0, "sku"] == "A"


def _make_multi_sheet_xlsx() -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Stores"
    ws1.append(["store_id", "sales"])
    ws1.append(["1", "100"])
    ws2 = wb.create_sheet("Archive")
    ws2.append(["store_id", "sales"])
    ws2.append(["9", "5"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_list_excel_sheets():
    from tabular_loader import list_excel_sheets

    sheets = list_excel_sheets(_make_multi_sheet_xlsx(), "multi.xlsx")
    assert sheets == ["Stores", "Archive"]


def test_load_excel_specific_sheet():
    df = load_tabular(_make_multi_sheet_xlsx(), "multi.xlsx", sheet_name="Archive")
    assert len(df) == 1
    assert df.loc[0, "store_id"] == "9"


def test_load_parquet():
    pytest.importorskip("pyarrow")
    import pandas as pd

    buf = io.BytesIO()
    pd.DataFrame({"x": ["a", "b"], "y": ["1", "2"]}).to_parquet(buf, index=False)
    df = load_tabular(buf.getvalue(), "data.parquet")
    assert list(df.columns) == ["x", "y"]
    assert len(df) == 2


def test_dedupe_columns_on_duplicate_headers():
    df = pd.DataFrame([[1, 2, 3]], columns=["A", "A", "B"])
    out = normalize_dataframe(df)
    assert list(out.columns) == ["A", "A__1", "B"]


def test_normalize_strips_excel_unnamed_empty_columns():
    df = pd.DataFrame({"a": [1, 2], "Unnamed: 1": [pd.NA, pd.NA]})
    out = normalize_dataframe(df)
    assert list(out.columns) == ["a"]


def test_empty_csv_rejected():
    with pytest.raises(ValueError, match="no rows|Empty"):
        from tabular_loader import load_tabular

        load_tabular(b"name\n", "headers_only.csv")
